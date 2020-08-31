from __future__ import print_function
import pickle
import os.path
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
import openpyxl
import datetime
import json


#Client ID: 290885468943-019j3pbpljoud8q5tst43hm3pirj2orc.apps.googleusercontent.com
#Client Secret: noix_yDK4jPYFrtyew76pIk9


def get_data_from_excel(pwd):
    """
    :param pwd: path for excel file
    :return: list of dates when i have billing
    """
    list_of_dates = list()
    wb = openpyxl.load_workbook(pwd, data_only=True)
    ws = wb['Taski 2019']
#    print(ws['A5'].fill.start_color.index)

    for cell in ws.iter_cols(min_row=13, max_row=13, min_col=1):
        if cell[0].fill.start_color.index == 'FFF381CA':
            list_of_dates.append(ws.cell(2, cell[0].column).value)
#            print(cell[0].fill.start_color.index)
    return list_of_dates


def grp_dates(list_of_dates):
    """
    :param list_of_dates: list of dates returned by get_data_from_excel()
    :return: list of tuples, tuple[0] 1st day of billing, tuple[1] last day of billing
    """
    calculated_events = list()
    event = (list_of_dates[0],)
    for i in range(1, len(list_of_dates)):
        #print(list_of_dates[i]-list_of_dates[i-1])
        if list_of_dates[i]-list_of_dates[i-1] > datetime.timedelta(days=1):
            event = event + (list_of_dates[i-1],)
            calculated_events.append(event)
            event = (list_of_dates[i],)

    return calculated_events


def create_google_event(list_of_events):
    """
    Function creates event data for google API
    :param list_of_events: returned by grp_dates()
    :return: list of dicts with events for google API
    """
    google_events = list()

    for event in list_of_events:
        google_events.append(
            {
                'summary': 'Dyżur Bilingowy',
                'start': {
                    'dateTime': event[0].strftime("%Y-%m-%dT%H:%M:%S+02:00"),
                },
                'end': {
                    'dateTime': event[1].replace(hour=23, minute=59).strftime("%Y-%m-%dT%H:%M:%S+02:00"),
                },
                'reminders': {
                    'useDefault': False,
                    'overrides': [
                        {'method': 'popup', 'minutes': 48 * 60},
                        {'method': 'popup', 'minutes': 24 * 60},
                    ],
                },
            })
#    print(google_events)
    return google_events


def add_google_event(google_events):
    SCOPES = ['https://www.googleapis.com/auth/calendar']
    creds = None
    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            creds = pickle.load(token)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('token.pickle', 'wb') as token:
            pickle.dump(creds, token)

    service = build('calendar', 'v3', credentials=creds)

    clear_calendar(service, '7i988gua5goaut69vvk8k13lvs@group.calendar.google.com')
    for event in google_events:
        event = service.events()\
            .insert(calendarId='7i988gua5goaut69vvk8k13lvs@group.calendar.google.com', body=event)\
            .execute()
        print('Event created: %s' % (event.get('htmlLink')))


def clear_calendar(service, calendar_id):
    page_token = None
    while True:
        events = service.events().list(calendarId=calendar_id, pageToken=page_token).execute()
        for event in events['items']:
            service.events().delete(calendarId=calendar_id, eventId=event['id']).execute()
        page_token = events.get('nextPageToken')
        if not page_token:
            print("Calendar: " + calendar_id + " is cleared")
            break


if __name__ == '__main__':
    dates = get_data_from_excel('60_Billing Operations Schedules/support_team_schedule_2020.xlsx')
    events = grp_dates(dates)
    gevents = create_google_event(events)
    add_google_event(gevents)